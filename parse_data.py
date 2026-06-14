"""parse_data.py — Parse VSTu bi-weekly xlsx report from raw data sheets."""
import re
from io import BytesIO
import pandas as pd
import openpyxl

THB_TO_VND = 830

BRANDS = [
    'WEARETHEPRIVATE', 'TRAPPER CLUB', 'FVNXYTHINGS', 'IAMSAIGON',
    'AFTERPARTY', 'FLORALPUNK', 'DEARJOSÉ', 'DEARJOSE', 'BLACKDRP',
    'BLACKORP', 'HIGHCHIC', 'PARADOX', 'MOIDIEN', 'MOIDEN', 'CAOSTU',
    'KANTAN', 'FNOS', 'QEM',
]

_PREV_KEYWORDS = ('prev', ' pre', 'apr', 'may', 'old', 'last', 'trước', 'truoc', 'march', 'mar',
                  'jan', 'feb', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec')


# ── helpers ──────────────────────────────────────────────────────────────────

def _n(v, default=0):
    if v is None:
        return default
    try:
        return float(str(v).replace(',', '').replace('%', '').strip())
    except (ValueError, TypeError):
        return default


def _s(v):
    return str(v).strip() if v is not None else ''


def extract_brand(name):
    upper = str(name).upper()
    for b in BRANDS:
        if b.upper() in upper:
            b_norm = b.replace('DEARJOSÉ', 'DEARJOSE').replace('BLACKORP', 'BLACKDRP').replace('MOIDEN', 'MOIDIEN')
            return b_norm
    return 'OTHER'


def classify_campaign(campaign_name):
    c = str(campaign_name).lower()
    if 'retarget' in c:                          return 'Retargeting'
    if 'catalogsale' in c or 'catalog' in c:     return 'Catalog Sale'
    if 'reach' in c:                             return 'Reach'
    if 'engagement' in c or 'engag' in c:        return 'Engagement'
    if 'visit profile' in c or 'profile' in c:   return 'Profile Visit'
    return 'Other'


def extract_format(ad_name):
    lower = str(ad_name).lower()
    if 'carousel' in lower:    return 'Carousel'
    if 'video' in lower:       return 'Video'
    if 'photo' in lower:       return 'Photo'
    return 'Other'


def extract_promotion(ad_name):
    """Extract promotion name from ad name — last segment after ' - ', strip copy/version suffixes."""
    parts = str(ad_name).split(' - ')
    if len(parts) >= 4:
        promo = parts[-1].strip()
        promo = re.sub(r'\s*[–\-]\s*(copy|v\d+)\s*$', '', promo, flags=re.IGNORECASE).strip()
        return promo if promo else 'Other'
    return 'Other'


def _is_prev_sheet(name):
    n = name.lower()
    return any(kw in n for kw in _PREV_KEYWORDS)


# ── Sheet detection ───────────────────────────────────────────────────────────

def _detect_sheets(wb):
    """
    Auto-detect sheet roles by content, not by name.
    Returns dict: {role: sheet_name}
    """
    roles = {
        'media_plan':      None,
        'fb_current':      None,
        'fb_prev':         None,
        'shopee_current':  None,
        'shopee_prev':     None,
        'tiktok_current':  None,
        'tiktok_prev':     None,
    }

    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = list(ws.iter_rows(min_row=1, max_row=12, values_only=True))
        flat = ' '.join(str(v).lower() for r in rows for v in r if v)

        # Media Plan — must have channel+budget structure; prefer sheet with GMV column
        is_media_plan = (('media plan' in flat and 'channel' in flat and 'budget' in flat) or
                         ('budget' in flat and 'channel' in flat and 'kpi' in flat))
        if is_media_plan:
            if roles['media_plan'] is None or 'gmv' in flat:
                roles['media_plan'] = sname
            continue

        # FB raw: has "ad name" and "amount spent"
        if 'ad name' in flat and 'amount spent' in flat:
            if _is_prev_sheet(sname):
                if roles['fb_prev'] is None:
                    roles['fb_prev'] = sname
            else:
                if roles['fb_current'] is None:
                    roles['fb_current'] = sname
            continue

        # TikTok raw: "ad name" + "cost" + ("video" or "tiktok" in sheet name), NOT "amount spent"
        if ('ad name' in flat and 'cost' in flat and 'amount spent' not in flat and
                ('video' in flat or 'tiktok' in sname.lower())):
            if _is_prev_sheet(sname):
                if roles['tiktok_prev'] is None:
                    roles['tiktok_prev'] = sname
            else:
                if roles['tiktok_current'] is None:
                    roles['tiktok_current'] = sname
            continue

        # Shopee: has "gmv" + "expense" (both product & campaign sheets)
        if 'gmv' in flat and 'expense' in flat and ('impression' in flat or 'clicks' in flat):
            if _is_prev_sheet(sname):
                if roles['shopee_prev'] is None:
                    roles['shopee_prev'] = sname
            else:
                if roles['shopee_current'] is None:
                    roles['shopee_current'] = sname

    return roles


# ── Media Plan parser ─────────────────────────────────────────────────────────

def parse_media_plan(wb, sheet_name):
    if not sheet_name:
        return {'total_budget': 0, 'timeline': '', 'channels': {}}

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    result = {'total_budget': 0, 'timeline': '', 'channels': {}}

    for row in rows[:10]:
        for i, v in enumerate(row):
            if v and 'budget' in str(v).lower() and i < 5:
                nxt = row[i + 1] if i + 1 < len(row) else None
                try:
                    result['total_budget'] = float(str(nxt).replace(',', ''))
                except Exception:
                    pass
            if v and 'timeline' in str(v).lower() and i < 5:
                tl = row[i + 1] if i + 1 < len(row) else None
                if tl:
                    result['timeline'] = str(tl).strip()

    # Channel rows: find header row then read data rows
    header_row = None
    for i, row in enumerate(rows):
        flat = ' '.join(str(v).lower() for v in row if v)
        if 'channel' in flat and 'budget' in flat:
            header_row = i
            break

    if header_row is not None:
        channel_map = {
            'reach':               'IG Reach',
            'engagement':          'IG Engagement',
            'profile visit':       'FB Profile Visit',
            'catalog sale':        'FB Catalog Sale',
            'dymanic retargeting': 'FB Retargeting',
            'dynamic retargeting': 'FB Retargeting',
            'gmv max':             'Shopee GMV Max',
            'shopee':              'Shopee GMV Max',
            'tiktok':              'TikTok',
        }

        # Detect columns from BOTH header row and sub-header row (row below)
        hdr = [str(v).lower().strip() if v else '' for v in rows[header_row]]
        sub = [str(v).lower().strip() if v else ''
               for v in rows[header_row + 1]] if header_row + 1 < len(rows) else []
        combined = [(sub[i] if i < len(sub) and sub[i] else hdr[i])
                    for i in range(max(len(hdr), len(sub)))]

        def _find_col(keywords, default):
            for i, h in enumerate(combined):
                if any(kw in h for kw in keywords):
                    return i
            return default

        budget_col = _find_col(['budget', 'งบ'], 6)
        kpi_col    = _find_col(['inventory buying', 'inventory'], 8)
        gmv_col    = _find_col(['gmv'], 22)
        pdp_col    = _find_col(['pdp/atc', 'pdp'], 21)

        # Scan all rows below header until 3 consecutive empty rows.
        # Only check col 1 (channel) + col 2 (format) to avoid matching
        # keywords inside long "Targeting Audience" description text.
        # Skip rows where col 2 is None — these are section headers / totals.
        consecutive_empty = 0
        for row in rows[header_row + 1:]:
            if not any(v for v in row):
                consecutive_empty += 1
                if consecutive_empty >= 3:
                    break
                continue
            consecutive_empty = 0

            # Skip section header / total rows (no format in col 2)
            if len(row) <= 2 or row[2] is None:
                continue

            check_text = ' '.join(
                str(row[i]).lower().strip()
                for i in [1, 2]
                if i < len(row) and row[i]
            )
            for key, label in channel_map.items():
                if key in check_text:
                    budget = _n(row[budget_col]) if len(row) > budget_col else 0
                    gmv    = _n(row[gmv_col])    if len(row) > gmv_col    else 0
                    if label in ('FB Catalog Sale', 'FB Retargeting'):
                        kpi = _n(row[pdp_col]) if len(row) > pdp_col else _n(row[kpi_col]) if len(row) > kpi_col else 0
                    else:
                        kpi = _n(row[kpi_col]) if len(row) > kpi_col else 0
                    if budget > 0:
                        kpi_orders = _n(row[kpi_col]) if len(row) > kpi_col else 0
                        kpi_funnel = _n(row[pdp_col]) if len(row) > pdp_col else kpi_orders
                        existing = result['channels'].get(label)
                        if not existing or (existing['gmv'] == 0 and gmv > 0):
                            result['channels'][label] = {
                                'budget':     budget,
                                'kpi':        kpi_orders,
                                'kpi_funnel': kpi_funnel,
                                'gmv':        gmv,
                            }
                    break

    # Fallback: compute total_budget from channel sum if not parsed directly
    if result['total_budget'] == 0 and result['channels']:
        result['total_budget'] = sum(ch['budget'] for ch in result['channels'].values())

    return result


# ── FB raw parser ─────────────────────────────────────────────────────────────

def _parse_fb_sheet(wb, sheet_name):
    """Parse a FB raw sheet into a cleaned DataFrame."""
    if not sheet_name or sheet_name not in wb.sheetnames:
        return pd.DataFrame()

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    # Find header row (has "ad name" and "amount spent")
    header_idx = None
    for i, row in enumerate(rows[:10]):
        flat = ' '.join(str(v).lower() for v in row if v)
        if 'ad name' in flat and 'amount spent' in flat:
            header_idx = i
            break
    if header_idx is None:
        return pd.DataFrame()

    headers = [str(h).strip() if h else f'col_{i}' for i, h in enumerate(rows[header_idx])]
    data = [list(r) for r in rows[header_idx + 1:] if any(v is not None for v in r)]
    df = pd.DataFrame(data, columns=headers)

    col_map = {}
    for col in df.columns:
        cl = col.lower().strip()
        if cl == 'ad name':                                  col_map[col] = 'Ad Name'
        elif cl == 'campaign name':                          col_map[col] = 'Campaign'
        elif cl == 'ad set name':                            col_map[col] = 'Ad Set'
        elif cl == 'ad delivery':                            col_map[col] = 'Status'
        elif 'amount spent' in cl:                           col_map[col] = 'Spend'
        elif cl == 'impressions':                            col_map[col] = 'Impressions'
        elif cl == 'reach':                                  col_map[col] = 'Reach'
        elif cl == 'clicks (all)':                           col_map[col] = 'Clicks'
        elif cl == 'ctr (all)':                              col_map[col] = 'CTR'
        elif cl == 'purchases':                              col_map[col] = 'Purchases'
        elif 'purchases conversion value' in cl:             col_map[col] = 'Revenue'
        elif 'purchase roas' in cl:                          col_map[col] = 'ROAS'
        elif cl == 'adds to cart':                           col_map[col] = 'ATC'
        elif cl == 'checkouts initiated':                    col_map[col] = 'Checkouts'
        elif 'thruplay' in cl and 'cost' not in cl:          col_map[col] = 'ThruPlays'
        elif cl == 'content views':                          col_map[col] = 'Content Views'
        elif 'instagram follows' in cl:                      col_map[col] = 'IG Follows'
        elif cl == 'result indicator':                       col_map[col] = 'Result Type'
        elif cl == 'results':                                col_map[col] = 'Results'

    df = df.rename(columns=col_map)

    num_cols = ['Spend', 'Impressions', 'Reach', 'Clicks', 'Purchases', 'Revenue',
                'ROAS', 'ATC', 'Checkouts', 'ThruPlays', 'Content Views',
                'Results', 'IG Follows', 'CTR']
    for col in num_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    if 'Ad Name' in df.columns:
        df['Brand']      = df['Ad Name'].apply(extract_brand)
        df['Format']     = df['Ad Name'].apply(extract_format)
        df['Promotion']  = df['Ad Name'].apply(extract_promotion)
    if 'Campaign' in df.columns:
        df['Camp Type']  = df['Campaign'].apply(classify_campaign)

    return df


# ── TikTok raw parser ────────────────────────────────────────────────────────

def _parse_tiktok_sheet(wb, sheet_name):
    """Parse TikTok Ads Manager export into a cleaned DataFrame."""
    if not sheet_name or sheet_name not in wb.sheetnames:
        return pd.DataFrame()

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    header_idx = None
    for i, row in enumerate(rows[:10]):
        flat = ' '.join(str(v).lower() for v in row if v)
        if 'ad name' in flat and ('cost' in flat or 'spend' in flat):
            header_idx = i
            break
    if header_idx is None:
        return pd.DataFrame()

    # Dedup headers to avoid duplicate column issues
    raw_headers = [str(h).strip() if h else f'col_{i}' for i, h in enumerate(rows[header_idx])]
    seen = {}
    headers = []
    for h in raw_headers:
        if h in seen:
            seen[h] += 1
            headers.append(f'{h}_{seen[h]}')
        else:
            seen[h] = 0
            headers.append(h)

    data = [list(r) for r in rows[header_idx + 1:] if any(v is not None for v in r)]
    if not data:
        return pd.DataFrame()
    df = pd.DataFrame(data, columns=headers)

    col_map = {}
    mapped = set()
    for col in df.columns:
        cl = col.lower().strip()
        if 'ad name' in cl and 'name' not in mapped:
            col_map[col] = 'Ad Name';      mapped.add('name')
        elif 'campaign name' in cl and 'campaign' not in mapped:
            col_map[col] = 'Campaign';     mapped.add('campaign')
        elif 'ad group' in cl and 'adgroup' not in mapped:
            col_map[col] = 'Ad Group';     mapped.add('adgroup')
        elif cl in ('impressions', 'impression') and 'impressions' not in mapped:
            col_map[col] = 'Impressions';  mapped.add('impressions')
        elif cl in ('clicks', 'click') and 'clicks' not in mapped:
            col_map[col] = 'Clicks';       mapped.add('clicks')
        elif cl == 'ctr' and 'ctr' not in mapped:
            col_map[col] = 'CTR';          mapped.add('ctr')
        elif cl in ('cost', 'spend', 'total cost') and 'spend' not in mapped:
            col_map[col] = 'Spend';        mapped.add('spend')
        elif ('video view' in cl or cl in ('vv', '2-second video views', '6-second video views')) and 'vv' not in mapped:
            col_map[col] = 'VideoViews';   mapped.add('vv')
        elif cl == 'reach' and 'reach' not in mapped:
            col_map[col] = 'Reach';        mapped.add('reach')
        elif cl == 'frequency' and 'frequency' not in mapped:
            col_map[col] = 'Frequency';    mapped.add('frequency')
        elif 'conversion' in cl and 'cost' not in cl and 'conversions' not in mapped:
            col_map[col] = 'Conversions';  mapped.add('conversions')
        elif ('campaign type' in cl or 'objective' in cl) and 'camptype' not in mapped:
            col_map[col] = 'Camp Type';    mapped.add('camptype')

    df = df.rename(columns=col_map)

    for col in ['Spend', 'Impressions', 'Clicks', 'VideoViews', 'Reach', 'Frequency', 'Conversions']:
        if col in df.columns:
            s = df[col]
            if isinstance(s, pd.DataFrame):
                s = s.iloc[:, 0]
            df[col] = pd.to_numeric(s, errors='coerce').fillna(0)

    if 'Ad Name' in df.columns:
        df['Brand']  = df['Ad Name'].apply(extract_brand)
        df['Format'] = df['Ad Name'].apply(extract_format)
    if 'Camp Type' not in df.columns and 'Campaign' in df.columns:
        df['Camp Type'] = df['Campaign'].apply(classify_campaign)

    return df


def _agg_tiktok(df):
    """Aggregate TikTok raw data for the TikTok tab."""
    if df.empty:
        return {'total': {}, 'by_format': [], 'top_ads': []}

    def _s(col):
        return float(df[col].sum()) if col in df.columns else 0.0

    total = {
        'spend':       _s('Spend'),
        'impressions': _s('Impressions'),
        'video_views': _s('VideoViews'),
        'reach':       _s('Reach'),
        'clicks':      _s('Clicks'),
        'conversions': _s('Conversions'),
    }
    total['vvr'] = total['video_views'] / total['impressions'] if total['impressions'] > 0 else 0
    total['ctr'] = total['clicks'] / total['impressions'] if total['impressions'] > 0 else 0

    # By Format/Camp Type
    by_format = []
    group_col = 'Camp Type' if 'Camp Type' in df.columns else ('Format' if 'Format' in df.columns else None)
    if group_col:
        for gtype, gdf in df.groupby(group_col):
            if not gtype or str(gtype).strip() in ('', 'nan', 'Other'):
                continue
            imp = gdf['Impressions'].sum() if 'Impressions' in gdf else 0
            vv  = gdf['VideoViews'].sum()  if 'VideoViews'  in gdf else 0
            by_format.append({
                'type':        str(gtype),
                'spend':       gdf['Spend'].sum()       if 'Spend'       in gdf else 0,
                'impressions': imp,
                'video_views': vv,
                'reach':       gdf['Reach'].sum()       if 'Reach'       in gdf else 0,
                'clicks':      gdf['Clicks'].sum()      if 'Clicks'      in gdf else 0,
                'vvr':         vv / imp if imp > 0 else 0,
            })
        by_format.sort(key=lambda x: x['spend'], reverse=True)

    # Top ads by VideoViews (fallback to Impressions)
    sort_col = 'VideoViews' if 'VideoViews' in df.columns else 'Impressions'
    top_df   = df.sort_values(sort_col, ascending=False).head(10) if sort_col in df.columns else df.head(10)
    top_ads  = top_df.to_dict('records')

    return {'total': total, 'by_format': by_format, 'top_ads': top_ads}


# ── Shopee raw parsers ────────────────────────────────────────────────────────

def _parse_shopee_sheet(wb, sheet_name):
    """
    Parse Shopee raw sheet (campaign-level hoặc product-level đều được).
    Returns cleaned DataFrame (VNĐ) với columns chuẩn hóa.
    """
    if not sheet_name or sheet_name not in wb.sheetnames:
        return pd.DataFrame()

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    # Find header row: has 'gmv' and 'expense'
    header_idx = None
    for i, row in enumerate(rows[:15]):
        flat = ' '.join(str(v).lower() for v in row if v)
        if 'gmv' in flat and 'expense' in flat and ('impression' in flat or 'clicks' in flat):
            header_idx = i
            break
    if header_idx is None:
        return pd.DataFrame()

    # De-duplicate headers before building DataFrame
    raw_headers = [str(h).strip() if h else f'col_{i}' for i, h in enumerate(rows[header_idx])]
    seen = {}
    headers = []
    for h in raw_headers:
        if h in seen:
            seen[h] += 1
            headers.append(f'{h}_{seen[h]}')
        else:
            seen[h] = 0
            headers.append(h)

    data = [list(r) for r in rows[header_idx + 1:] if any(v is not None for v in r)]
    if not data:
        return pd.DataFrame()

    df = pd.DataFrame(data, columns=headers)

    # Map to canonical names — use FIRST occurrence of each concept
    col_map = {}
    mapped = set()

    for col in df.columns:
        cl = col.lower().strip().rstrip('_0123456789')
        if ('ad / product name' in cl or 'ad name' == cl) and 'name' not in mapped:
            col_map[col] = 'name'; mapped.add('name')
        elif cl in ('impression', 'impressions') and 'impressions' not in mapped:
            col_map[col] = 'impressions'; mapped.add('impressions')
        elif cl == 'clicks' and 'clicks' not in mapped:
            col_map[col] = 'clicks'; mapped.add('clicks')
        elif cl == 'ctr' and 'ctr' not in mapped:
            col_map[col] = 'ctr'; mapped.add('ctr')
        elif cl == 'items sold' and 'orders' not in mapped:
            col_map[col] = 'orders'; mapped.add('orders')
        elif cl == 'conversions' and 'orders' not in mapped:
            col_map[col] = 'orders'; mapped.add('orders')
        elif cl == 'gmv' and 'gmv_thb' not in mapped:
            col_map[col] = 'gmv_thb'; mapped.add('gmv_thb')
        elif cl == 'expense' and 'spend_thb' not in mapped:
            col_map[col] = 'spend_thb'; mapped.add('spend_thb')
        elif cl == 'roas' and 'roas' not in mapped:
            col_map[col] = 'roas'; mapped.add('roas')
        elif cl in ('status', 'ad status') and 'status' not in mapped:
            col_map[col] = 'status'; mapped.add('status')

    df = df.rename(columns=col_map)

    # Parse numeric columns
    for col in ['impressions', 'clicks', 'orders', 'gmv_thb', 'spend_thb', 'roas']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    if 'ctr' in df.columns:
        df['ctr'] = df['ctr'].apply(
            lambda x: _n(str(x).replace('%', '')) / 100
            if isinstance(x, str) and '%' in str(x) else _n(x)
        )

    # Convert THB → VNĐ
    if 'gmv_thb'   in df.columns: df['gmv']   = df['gmv_thb']   * THB_TO_VND
    if 'spend_thb' in df.columns: df['spend'] = df['spend_thb'] * THB_TO_VND

    # Recompute ROAS
    if 'gmv' in df.columns and 'spend' in df.columns:
        df['roas'] = df.apply(lambda r: r['gmv'] / r['spend'] if r['spend'] > 0 else 0, axis=1)

    # Drop empty / sequence-number rows
    if 'name' in df.columns:
        df = df[df['name'].apply(lambda x: bool(_s(x)) and not str(x).replace('.', '').isdigit())]
        df['brand'] = df['name'].apply(extract_brand)

    return df.reset_index(drop=True)


# ── Aggregate from raw ────────────────────────────────────────────────────────

def _agg_branding(df_fb):
    """Compute branding summary from raw FB DataFrame."""
    if df_fb.empty or 'Camp Type' not in df_fb.columns:
        return {'summary': [], 'top_ads': []}

    summary = []
    for camp_type, label in [('Reach', 'Reach'), ('Engagement', 'Engagement'), ('Profile Visit', 'Profile Visit')]:
        sub = df_fb[df_fb['Camp Type'] == camp_type]
        reach = sub['Reach'].sum() if 'Reach' in sub.columns else 0
        imps  = sub['Impressions'].sum() if 'Impressions' in sub.columns else 0
        summary.append({
            'type':        label,
            'spend':       sub['Spend'].sum() if 'Spend' in sub.columns else 0,
            'impressions': imps,
            'reach':       reach,
            'engagement':  sub['Results'].sum() if camp_type == 'Engagement' and 'Results' in sub.columns else 0,
            'followers':   sub['IG Follows'].sum() if 'IG Follows' in sub.columns else 0,
            'clicks':      sub['Clicks'].sum() if 'Clicks' in sub.columns else 0,
            'video_views': sub['ThruPlays'].sum() if 'ThruPlays' in sub.columns else 0,
            'frequency':   imps / reach if reach > 0 else 0,
        })

    sub_all = df_fb[df_fb['Camp Type'].isin(['Reach', 'Engagement', 'Profile Visit'])]
    reach_t = sub_all['Reach'].sum() if 'Reach' in sub_all.columns else 0
    imps_t  = sub_all['Impressions'].sum() if 'Impressions' in sub_all.columns else 0
    summary.append({
        'type':        'Total',
        'spend':       sub_all['Spend'].sum() if 'Spend' in sub_all.columns else 0,
        'impressions': imps_t,
        'reach':       reach_t,
        'engagement':  df_fb[df_fb['Camp Type'] == 'Engagement']['Results'].sum()
                       if 'Results' in df_fb.columns else 0,
        'followers':   sub_all['IG Follows'].sum() if 'IG Follows' in sub_all.columns else 0,
        'clicks':      sub_all['Clicks'].sum() if 'Clicks' in sub_all.columns else 0,
        'video_views': sub_all['ThruPlays'].sum() if 'ThruPlays' in sub_all.columns else 0,
        'frequency':   imps_t / reach_t if reach_t > 0 else 0,
    })

    # Top 3 ads by engagement
    top_ads = []
    eng_df = df_fb[df_fb['Camp Type'] == 'Engagement'].copy()
    if not eng_df.empty and 'Ad Name' in eng_df.columns:
        agg_cols = {c: 'sum' for c in ['Results', 'ThruPlays', 'Impressions', 'Spend'] if c in eng_df.columns}
        eng_agg = eng_df.groupby('Ad Name').agg(agg_cols).reset_index()
        eng_agg['vtr'] = eng_agg.apply(
            lambda r: r.get('ThruPlays', 0) / r['Impressions'] if r.get('Impressions', 0) > 0 else 0, axis=1
        )
        for _, row in eng_agg.sort_values('Results', ascending=False).head(3).iterrows():
            top_ads.append({
                'name':       row['Ad Name'],
                'brand':      extract_brand(row['Ad Name']),
                'format':     extract_format(row['Ad Name']),
                'engagement': row.get('Results', 0),
                'vtr':        row.get('vtr', 0),
                'spend':      row.get('Spend', 0),
            })

    return {'summary': summary, 'top_ads': top_ads}


def _agg_conversion(df_fb):
    """Compute FB conversion funnel from raw FB DataFrame."""
    if df_fb.empty or 'Camp Type' not in df_fb.columns:
        return []

    result = []
    for camp_type, label in [('Catalog Sale', 'Catalog Sale'), ('Retargeting', 'Retargeting')]:
        sub = df_fb[df_fb['Camp Type'] == camp_type]
        spend     = sub['Spend'].sum() if 'Spend' in sub.columns else 0
        pdp_views = sub['Content Views'].sum() if 'Content Views' in sub.columns else 0
        atc       = sub['ATC'].sum() if 'ATC' in sub.columns else 0
        checkouts = sub['Checkouts'].sum() if 'Checkouts' in sub.columns else 0
        purchases = sub['Purchases'].sum() if 'Purchases' in sub.columns else 0
        revenue   = sub['Revenue'].sum() if 'Revenue' in sub.columns else 0
        imps      = sub['Impressions'].sum() if 'Impressions' in sub.columns else 0
        reach     = sub['Reach'].sum() if 'Reach' in sub.columns else 0
        roas      = revenue / spend if spend > 0 else 0
        cr        = purchases / pdp_views if pdp_views > 0 else 0
        result.append({
            'type':         label,
            'spend':        spend,
            'impressions':  imps,
            'reach':        reach,
            'ctr':          sub['CTR'].mean() if 'CTR' in sub.columns and len(sub) else 0,
            'pdp_views':    pdp_views,
            'a2c_rate':     atc / pdp_views if pdp_views > 0 else 0,
            'atc':          atc,
            'co_rate':      checkouts / atc if atc > 0 else 0,
            'checkouts':    checkouts,
            'pur_rate':     purchases / checkouts if checkouts > 0 else 0,
            'purchases':    purchases,
            'cr':           cr,
            'cost_per_pur': spend / purchases if purchases > 0 else 0,
            'gmv':          revenue,
            'roas':         roas,
        })

    # Total row
    sub_all = df_fb[df_fb['Camp Type'].isin(['Catalog Sale', 'Retargeting'])]
    total_spend = sub_all['Spend'].sum() if 'Spend' in sub_all.columns else 0
    total_pdp   = sub_all['Content Views'].sum() if 'Content Views' in sub_all.columns else 0
    total_atc   = sub_all['ATC'].sum() if 'ATC' in sub_all.columns else 0
    total_co    = sub_all['Checkouts'].sum() if 'Checkouts' in sub_all.columns else 0
    total_pur   = sub_all['Purchases'].sum() if 'Purchases' in sub_all.columns else 0
    total_rev   = sub_all['Revenue'].sum() if 'Revenue' in sub_all.columns else 0
    result.append({
        'type':         'Total',
        'spend':        total_spend,
        'impressions':  sub_all['Impressions'].sum() if 'Impressions' in sub_all.columns else 0,
        'reach':        sub_all['Reach'].sum() if 'Reach' in sub_all.columns else 0,
        'pdp_views':    total_pdp,
        'a2c_rate':     total_atc / total_pdp if total_pdp > 0 else 0,
        'atc':          total_atc,
        'co_rate':      total_co / total_atc if total_atc > 0 else 0,
        'checkouts':    total_co,
        'pur_rate':     total_pur / total_co if total_co > 0 else 0,
        'purchases':    total_pur,
        'cr':           total_pur / total_pdp if total_pdp > 0 else 0,
        'cost_per_pur': total_spend / total_pur if total_pur > 0 else 0,
        'gmv':          total_rev,
        'roas':         total_rev / total_spend if total_spend > 0 else 0,
    })
    return result


def _agg_shopee_overall(df):
    if df.empty:
        return {'spend': 0, 'orders': 0, 'gmv': 0, 'roas': 0, 'clicks': 0, 'impressions': 0}
    spend  = df['spend'].sum()  if 'spend'  in df.columns else 0
    orders = df['orders'].sum() if 'orders' in df.columns else 0
    gmv    = df['gmv'].sum()    if 'gmv'    in df.columns else 0
    return {
        'spend':       spend,
        'orders':      orders,
        'gmv':         gmv,
        'roas':        gmv / spend if spend > 0 else 0,
        'clicks':      df['clicks'].sum()      if 'clicks'      in df.columns else 0,
        'impressions': df['impressions'].sum() if 'impressions' in df.columns else 0,
    }


def _agg_shopee_brands(df_prod):
    """Aggregate Shopee product data by brand."""
    if df_prod.empty or 'brand' not in df_prod.columns:
        return []
    agg = df_prod.groupby('brand').agg(
        orders=('orders', 'sum'),
        gmv=('gmv', 'sum'),
        spend=('spend', 'sum'),
    ).reset_index()
    agg['roas'] = agg.apply(lambda r: r['gmv'] / r['spend'] if r['spend'] > 0 else 0, axis=1)
    return agg.sort_values('gmv', ascending=False).to_dict('records')


def _agg_shopee_products(df, top_n=25):
    """Aggregate Shopee data by product/campaign name for MoM comparison."""
    if df.empty or 'name' not in df.columns:
        return []
    num_cols = [c for c in ['gmv', 'spend', 'orders', 'clicks', 'impressions'] if c in df.columns]
    if not num_cols:
        return []
    agg = df.groupby('name')[num_cols].sum().reset_index()
    if 'gmv' in agg.columns and 'spend' in agg.columns:
        agg['roas'] = agg.apply(lambda r: r['gmv'] / r['spend'] if r['spend'] > 0 else 0, axis=1)
    else:
        agg['roas'] = 0
    if 'brand' in df.columns:
        brand_map = df.groupby('name')['brand'].first()
        agg['brand'] = agg['name'].map(brand_map).fillna('OTHER')
    else:
        agg['brand'] = 'OTHER'
    return agg.sort_values('gmv', ascending=False).head(top_n).to_dict('records')


def _agg_sale_summary(df_fb, df_shopeeamp):
    """Compute overall sale performance (for Overview tab)."""
    result = {'channels': [], 'total': {}}

    fb_conv = df_fb[df_fb['Camp Type'].isin(['Catalog Sale', 'Retargeting'])] if not df_fb.empty else pd.DataFrame()
    for camp_type, label in [('Catalog Sale', 'FB Catalog Sale'), ('Retargeting', 'FB Retargeting')]:
        sub = fb_conv[fb_conv['Camp Type'] == camp_type] if not fb_conv.empty else pd.DataFrame()
        result['channels'].append({
            'channel': label,
            'spend':   sub['Spend'].sum() if not sub.empty and 'Spend' in sub.columns else 0,
            'orders':  sub['Purchases'].sum() if not sub.empty and 'Purchases' in sub.columns else 0,
            'gmv':     sub['Revenue'].sum() if not sub.empty and 'Revenue' in sub.columns else 0,
        })

    shopee_spend  = df_shopeeamp['spend'].sum()  if not df_shopeeamp.empty and 'spend'  in df_shopeeamp.columns else 0
    shopee_orders = df_shopeeamp['orders'].sum() if not df_shopeeamp.empty and 'orders' in df_shopeeamp.columns else 0
    shopee_gmv    = df_shopeeamp['gmv'].sum()    if not df_shopeeamp.empty and 'gmv'    in df_shopeeamp.columns else 0
    # Note: df_shopeeamp is the combined Shopee sheet (name col, not campaign col)
    result['channels'].append({
        'channel': 'Shopee', 'spend': shopee_spend,
        'orders': shopee_orders, 'gmv': shopee_gmv,
    })

    for ch in result['channels']:
        ch['roas'] = ch['gmv'] / ch['spend'] if ch['spend'] > 0 else 0

    conv_spend = sum(c['spend'] for c in result['channels'])
    total_gmv  = sum(c['gmv']   for c in result['channels'])

    # spend_all = all FB campaigns (branding + conversion) + Shopee
    fb_all_spend = df_fb['Spend'].sum() if not df_fb.empty and 'Spend' in df_fb.columns else 0
    spend_all    = fb_all_spend + shopee_spend

    result['total'] = {
        'spend':     conv_spend,
        'spend_all': spend_all,
        'orders':    sum(c['orders'] for c in result['channels']),
        'gmv':       total_gmv,
        'roas':      total_gmv / conv_spend if conv_spend > 0 else 0,
    }
    return result


def _mom_from_raw(df_fb_prev, df_shopee_prev):
    """Compute previous period totals for MoM comparison."""
    if df_fb_prev.empty and df_shopee_prev.empty:
        return {}

    fb_conv = df_fb_prev[df_fb_prev['Camp Type'].isin(['Catalog Sale', 'Retargeting'])] \
              if not df_fb_prev.empty and 'Camp Type' in df_fb_prev.columns else pd.DataFrame()
    channels = []
    for camp_type, label in [('Catalog Sale', 'FB Catalog Sale'), ('Retargeting', 'FB Retargeting')]:
        sub = fb_conv[fb_conv['Camp Type'] == camp_type] if not fb_conv.empty else pd.DataFrame()
        channels.append({
            'channel': label,
            'spend':  sub['Spend'].sum()    if not sub.empty and 'Spend'    in sub.columns else 0,
            'orders': sub['Purchases'].sum() if not sub.empty and 'Purchases' in sub.columns else 0,
            'gmv':    sub['Revenue'].sum()  if not sub.empty and 'Revenue'  in sub.columns else 0,
        })

    sp = df_shopee_prev['spend'].sum()  if not df_shopee_prev.empty and 'spend'  in df_shopee_prev.columns else 0
    or_ = df_shopee_prev['orders'].sum() if not df_shopee_prev.empty and 'orders' in df_shopee_prev.columns else 0
    gm = df_shopee_prev['gmv'].sum()    if not df_shopee_prev.empty and 'gmv'    in df_shopee_prev.columns else 0
    channels.append({'channel': 'Shopee', 'spend': sp, 'orders': or_, 'gmv': gm})

    for ch in channels:
        ch['roas'] = ch['gmv'] / ch['spend'] if ch['spend'] > 0 else 0

    conv_spend = sum(c['spend'] for c in channels)
    total_gmv  = sum(c['gmv']   for c in channels)

    fb_all_spend_prev = df_fb_prev['Spend'].sum() if not df_fb_prev.empty and 'Spend' in df_fb_prev.columns else 0
    spend_all_prev    = fb_all_spend_prev + sp

    total = {
        'spend':     conv_spend,
        'spend_all': spend_all_prev,
        'orders':    sum(c['orders'] for c in channels),
        'gmv':       total_gmv,
        'roas':      total_gmv / conv_spend if conv_spend > 0 else 0,
    }
    return {'channels': channels, 'total': total}


def _detect_date_range(df_fb):
    """Extract date range from FB raw data."""
    if df_fb.empty:
        return {'raw': ''}
    for col in df_fb.columns:
        if 'reporting starts' in col.lower() or col.lower() == 'reporting starts':
            vals = df_fb[col].dropna()
            if len(vals):
                start = str(vals.iloc[0])[:10]
                end   = str(df_fb[df_fb.columns[df_fb.columns.str.lower().str.contains('end')]
                                  .tolist()[0] if any(df_fb.columns.str.lower().str.contains('end')) else col]
                            .dropna().iloc[-1])[:10] if len(vals) > 1 else start
                return {'raw': f'{start} – {end}'}
    return {'raw': ''}


# ── Main entry point ─────────────────────────────────────────────────────────

def parse_all(file_bytes):
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    roles = _detect_sheets(wb)

    plan           = parse_media_plan(wb, roles['media_plan'])
    df_fb          = _parse_fb_sheet(wb, roles['fb_current'])
    df_fb_prev     = _parse_fb_sheet(wb, roles['fb_prev'])
    df_shopee      = _parse_shopee_sheet(wb, roles['shopee_current'])
    df_shopee_p    = df_shopee
    df_shopeep     = _parse_shopee_sheet(wb, roles['shopee_prev'])
    df_tiktok      = _parse_tiktok_sheet(wb, roles['tiktok_current'])
    df_tiktok_prev = _parse_tiktok_sheet(wb, roles['tiktok_prev'])

    branding   = _agg_branding(df_fb)
    fb_conv    = _agg_conversion(df_fb)
    sale       = _agg_sale_summary(df_fb, df_shopee)
    sale_prev  = _mom_from_raw(df_fb_prev, df_shopeep)
    tiktok     = _agg_tiktok(df_tiktok)

    # Shopee campaign list — each row is one ad/campaign
    shopee_campaigns = []
    if not df_shopee.empty and 'name' in df_shopee.columns:
        for _, row in df_shopee.iterrows():
            shopee_campaigns.append({
                'campaign':    str(row.get('name', '')),
                'impressions': row.get('impressions', 0),
                'clicks':      row.get('clicks', 0),
                'ctr':         row.get('ctr', 0),
                'orders':      row.get('orders', 0),
                'cr':          row.get('orders', 0) / row.get('clicks', 1) if row.get('clicks') else 0,
                'gmv':         row.get('gmv', 0),
                'spend':       row.get('spend', 0),
                'roas':        row.get('roas', 0),
            })

    shopee_brands        = _agg_shopee_brands(df_shopee_p)
    shopee_products_cur  = _agg_shopee_products(df_shopee)
    shopee_products_prev = _agg_shopee_products(df_shopeep)

    # MoM funnel rates from FB raw
    mom_funnel = {}
    if not df_fb.empty and not df_fb_prev.empty:
        def _funnel_rates(df):
            sub = df[df['Camp Type'].isin(['Catalog Sale', 'Retargeting'])] \
                  if 'Camp Type' in df.columns else pd.DataFrame()
            pdp = sub['Content Views'].sum() if 'Content Views' in sub.columns else 0
            atc = sub['ATC'].sum()           if 'ATC'           in sub.columns else 0
            co  = sub['Checkouts'].sum()     if 'Checkouts'     in sub.columns else 0
            pur = sub['Purchases'].sum()     if 'Purchases'     in sub.columns else 0
            return {
                'a2c':      atc / pdp if pdp > 0 else 0,
                'checkout': co  / atc if atc > 0 else 0,
                'purchase': pur / co  if co  > 0 else 0,
                'cr':       pur / pdp if pdp > 0 else 0,
            }
        mom_funnel = {
            't_current': _funnel_rates(df_fb),
            't_prev':    _funnel_rates(df_fb_prev),
        }

    return {
        'plan':     plan,
        'overview': {
            'sale_current': sale['channels'],
            'sale_prev':    sale_prev.get('channels', []),
            'total_current': sale['total'],
            'total_prev':    sale_prev.get('total', {}),
            'date_range':    '',
            'overall':       _build_overall_table(plan, sale, branding, fb_conv, tiktok),
        },
        'branding':   branding,
        'conversion': {
            'fb':              fb_conv,
            'shopee_overall':  _agg_shopee_overall(df_shopee),
            'mom':             mom_funnel,
            'shopee_campaigns':    shopee_campaigns,
            'shopee_brands':       shopee_brands,
            'shopee_products_cur':  shopee_products_cur,
            'shopee_products_prev': shopee_products_prev,
        },
        'tiktok':     tiktok,
        'tiktok_raw': df_tiktok,
        'fb_raw':     df_fb,
        'date_range': _detect_date_range(df_fb),
        '_roles':     roles,
        '_sheets':    wb.sheetnames,
    }


def _build_overall_table(plan, sale, branding, fb_conv, tiktok=None):
    """Build plan vs actual rows for Overview tab."""
    rows = []
    ch_plan = plan.get('channels', {})
    br_sum  = {s['type']: s for s in branding.get('summary', [])}
    fb_dict = {f['type']: f for f in fb_conv}

    mapping = [
        ('FB Reach',        'IG Reach',         br_sum.get('Reach', {}),         None),
        ('IG Engagement',   'IG Engagement',     br_sum.get('Engagement', {}),    None),
        ('FB Profile Visit','FB Profile Visit',  br_sum.get('Profile Visit', {}), None),
        ('FB Catalog Sale', 'FB Catalog Sale',   fb_dict.get('Catalog Sale', {}), fb_dict.get('Catalog Sale', {})),
        ('FB Retargeting',  'FB Retargeting',    fb_dict.get('Retargeting', {}),  fb_dict.get('Retargeting', {})),
        ('Shopee GMV Max',  'Shopee GMV Max',    {}, None),
        ('TikTok',          'TikTok',            {}, None),
    ]

    shopee_sale  = next((c for c in sale['channels'] if c['channel'] == 'Shopee'), {})
    tiktok_total = (tiktok or {}).get('total', {})
    handled_keys = set()

    for label, plan_key, actual_data, conv_data in mapping:
        p = ch_plan.get(plan_key, {})
        handled_keys.add(plan_key)
        if label == 'Shopee GMV Max':
            actual_spend = shopee_sale.get('spend', 0)
            actual_kpi   = shopee_sale.get('orders', 0)
            kpi_plan     = p.get('kpi', 0)
            roas_actual  = shopee_sale.get('roas', 0)
        elif label == 'TikTok':
            actual_spend = tiktok_total.get('spend', 0)
            actual_kpi   = tiktok_total.get('video_views', tiktok_total.get('impressions', 0))
            kpi_plan     = p.get('kpi', 0)
            roas_actual  = 0
        elif label in ('FB Catalog Sale', 'FB Retargeting'):
            actual_spend = actual_data.get('spend', 0)
            actual_kpi   = (actual_data.get('pdp_views', 0) + actual_data.get('atc', 0) +
                            actual_data.get('checkouts', 0) + actual_data.get('purchases', 0))
            kpi_plan     = p.get('kpi_funnel', p.get('kpi', 0))
            gmv          = actual_data.get('gmv', 0)
            roas_actual  = gmv / actual_spend if actual_spend > 0 else 0
        else:
            actual_spend = actual_data.get('spend', 0)
            actual_kpi   = actual_data.get('impressions', actual_data.get('reach', actual_data.get('clicks', 0)))
            kpi_plan     = p.get('kpi', 0)
            roas_actual  = 0

        rows.append({
            'channel':        label,
            'budget_plan':    p.get('budget', 0),
            'budget_actual':  actual_spend,
            'budget_pct':     actual_spend / p.get('budget', 1) if p.get('budget') else 0,
            'kpi_plan':       kpi_plan,
            'kpi_actual':     actual_kpi,
            'kpi_pct':        actual_kpi / kpi_plan if kpi_plan > 0 else 0,
            'roas_actual':    roas_actual,
        })

    # Channels có trong plan nhưng chưa có trong mapping cứng → thêm vào cuối (actual = 0)
    for plan_key, p in ch_plan.items():
        if plan_key not in handled_keys and p.get('budget', 0) > 0:
            rows.append({
                'channel':       plan_key,
                'budget_plan':   p.get('budget', 0),
                'budget_actual': 0,
                'budget_pct':    0,
                'kpi_plan':      p.get('kpi', 0),
                'kpi_actual':    0,
                'kpi_pct':       0,
                'roas_actual':   0,
            })

    return rows
